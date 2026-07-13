import datetime
import io
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, status, Header
from fastapi.responses import StreamingResponse
from app.jwt_config import verificar_token
from sqlalchemy.orm import Session
from app.config.database import get_db
from app.models.models import Setor, Usuario, Chamado
from app.controllers.auth import get_current_user, get_current_gestor
from app.controllers.chamado import calcular_sla_chamado, calcular_tempos_por_setor, serializar_chamado
from app.controllers.request import abrirChamado
from app.schemas.schemas import LoginSchema, SetorSchema, PedidoSchema, UsuarioCadastroSchema, normalizar_status
from app.schemas.schemas import AtualizarPerfilSchema, AlterarSenhaSchema
router = APIRouter()


def _usuario_por_authorization(db: Session, authorization: Optional[str]) -> Usuario:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Token de autenticação obrigatório")
    dados = verificar_token(authorization.split(" ", 1)[1])
    usuario = db.query(Usuario).filter(Usuario.id == dados["usuario_id"]).first()
    if usuario is None:
        raise HTTPException(status_code=401, detail="Usuário não encontrado")
    if usuario.ativo is False:
        raise HTTPException(status_code=403, detail="Usuário inativo")
    return usuario


def _exigir_admin_ou_bootstrap(db: Session, authorization: Optional[str]) -> None:
    if db.query(Usuario).count() == 0:
        return
    usuario = _usuario_por_authorization(db, authorization)
    if usuario.perfil != "Administrador":
        raise HTTPException(status_code=403, detail="Apenas administradores podem realizar esta ação")


def _status_equivalentes(valor: str) -> list[str]:
    status_normalizado = normalizar_status(valor)
    equivalencias = {
        "Aberto": ["Aberto"],
        "Em Atendimento": ["Em Atendimento", "Em Progresso", "Em Andamento"],
        "Pausado": ["Pausado"],
        "Concluído": ["Concluído", "Concluido", "Fechado", "Resolvido"],
    }
    return equivalencias.get(status_normalizado, [valor])


def _resposta_paginada(items: list, total: int, page: Optional[int], per_page: int):
    if page is None:
        return items
    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page),
    }


def _serializar_chamado_lista(c: Chamado, db: Session) -> dict:
    setor_res = db.query(Setor).filter(Setor.id == c.setor_responsavel_id).first()
    usuario_res = None
    if c.usuario_responsavel_id:
        usuario_res = db.query(Usuario).filter(Usuario.id == c.usuario_responsavel_id).first()
    return {
        "id": c.id,
        "titulo": c.titulo,
        "descricao": c.descricao,
        "status": normalizar_status(c.status),
        "prioridade": c.prioridade,
        "data_criacao": c.data_criacao.isoformat() if c.data_criacao else None,
        "data_atualizacao": c.data_atualizacao.isoformat() if c.data_atualizacao else None,
        "setor_responsavel": setor_res.nome if setor_res else "Não informado",
        "usuario_responsavel": usuario_res.nome if usuario_res else "Enviar para todos (Nenhum específico)",
        "usuario_responsavel_id": c.usuario_responsavel_id,
        "sla": calcular_sla_chamado(c),
        "data_fechamento": c.data_fechamento.isoformat() if c.data_fechamento else None,
        "avaliacao": {
            "nota": c.avaliacao_nota,
            "comentario": c.avaliacao_comentario,
            "data_avaliacao": c.data_avaliacao.isoformat() if c.data_avaliacao else None,
        } if c.avaliacao_nota else None,
    }


def _parse_data(valor: Optional[str], fim_do_dia: bool = False):
    if not valor:
        return None
    try:
        data = datetime.datetime.strptime(valor, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=422, detail="Datas devem estar no formato YYYY-MM-DD")
    if fim_do_dia:
        return data.replace(hour=23, minute=59, second=59, microsecond=999999)
    return data


def _linhas_relatorio(chamados: list[Chamado], db: Session) -> tuple[list[dict], dict]:
    ids_setores = {c.setor_responsavel_id for c in chamados} | {c.setor_solicitante_id for c in chamados}
    ids_usuarios = {c.usuario_solicitante_id for c in chamados}
    ids_usuarios.update(c.usuario_responsavel_id for c in chamados if c.usuario_responsavel_id)
    setores = db.query(Setor).filter(Setor.id.in_(ids_setores)).all() if ids_setores else []
    usuarios = db.query(Usuario).filter(Usuario.id.in_(ids_usuarios)).all() if ids_usuarios else []
    setores_map = {s.id: s for s in setores}
    usuarios_map = {u.id: u for u in usuarios}

    linhas = []
    tempos_resposta = []
    notas = []
    atrasados = 0

    for chamado in chamados:
        sla = calcular_sla_chamado(chamado)
        tempos_setor = calcular_tempos_por_setor(chamado, db)
        tempos_validos = [t["tempo_resposta_horas"] for t in tempos_setor if t.get("tempo_resposta_horas") is not None]
        tempo_resposta = round(sum(tempos_validos) / len(tempos_validos), 2) if tempos_validos else None
        if tempo_resposta is not None:
            tempos_resposta.append(tempo_resposta)
        if chamado.avaliacao_nota:
            notas.append(chamado.avaliacao_nota)
        if sla["estado"] == "atrasado":
            atrasados += 1

        setor_responsavel = setores_map.get(chamado.setor_responsavel_id)
        setor_solicitante = setores_map.get(chamado.setor_solicitante_id)
        solicitante = usuarios_map.get(chamado.usuario_solicitante_id)
        responsavel = usuarios_map.get(chamado.usuario_responsavel_id)
        linhas.append({
            "id": chamado.id,
            "titulo": chamado.titulo,
            "status": normalizar_status(chamado.status),
            "prioridade": chamado.prioridade,
            "setor_solicitante": setor_solicitante.nome if setor_solicitante else "-",
            "setor_responsavel": setor_responsavel.nome if setor_responsavel else "-",
            "solicitante": solicitante.nome if solicitante else "-",
            "responsavel": responsavel.nome if responsavel else "Sem responsável",
            "data_criacao": chamado.data_criacao.isoformat() if chamado.data_criacao else None,
            "data_fechamento": chamado.data_fechamento.isoformat() if chamado.data_fechamento else None,
            "sla_estado": sla["estado"],
            "sla_percentual": sla["percentual"],
            "tempo_resposta_horas": tempo_resposta,
            "avaliacao_nota": chamado.avaliacao_nota,
            "avaliacao_comentario": chamado.avaliacao_comentario,
        })

    resumo = {
        "total": len(chamados),
        "abertos": sum(1 for c in chamados if normalizar_status(c.status) == "Aberto"),
        "em_atendimento": sum(1 for c in chamados if normalizar_status(c.status) == "Em Atendimento"),
        "pausados": sum(1 for c in chamados if normalizar_status(c.status) == "Pausado"),
        "concluidos": sum(1 for c in chamados if normalizar_status(c.status) == "Concluído"),
        "atrasados": atrasados,
        "tempo_medio_resposta_horas": round(sum(tempos_resposta) / len(tempos_resposta), 2) if tempos_resposta else None,
        "media_avaliacao": round(sum(notas) / len(notas), 2) if notas else None,
    }
    return linhas, resumo


def _resposta_xlsx(linhas: list[dict], resumo: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Chamados"
    headers = [
        "ID", "Titulo", "Status", "Prioridade", "Setor Solicitante", "Setor Responsavel",
        "Solicitante", "Responsavel", "Criacao", "Fechamento", "SLA", "% SLA",
        "Tempo Resposta (h)", "Nota", "Comentario Avaliacao"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for linha in linhas:
        ws.append([
            linha["id"], linha["titulo"], linha["status"], linha["prioridade"],
            linha["setor_solicitante"], linha["setor_responsavel"], linha["solicitante"],
            linha["responsavel"], linha["data_criacao"], linha["data_fechamento"],
            linha["sla_estado"], linha["sla_percentual"], linha["tempo_resposta_horas"],
            linha["avaliacao_nota"], linha["avaliacao_comentario"],
        ])
    ws_resumo = wb.create_sheet("Resumo")
    for chave, valor in resumo.items():
        ws_resumo.append([chave, valor])
    arquivo = io.BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return StreamingResponse(
        arquivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="relatorio-chamados.xlsx"'},
    )


def _resposta_pdf(linhas: list[dict], resumo: dict):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet

    arquivo = io.BytesIO()
    doc = SimpleDocTemplate(arquivo, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Relatório Gerencial de Chamados", styles["Title"]), Spacer(1, 12)]
    resumo_texto = " | ".join(f"{k}: {v if v is not None else '-'}" for k, v in resumo.items())
    elementos.append(Paragraph(resumo_texto, styles["Normal"]))
    elementos.append(Spacer(1, 12))
    dados = [["ID", "Título", "Status", "Setor", "SLA", "Resp. h", "Nota"]]
    for linha in linhas[:80]:
        dados.append([
            linha["id"],
            linha["titulo"][:42],
            linha["status"],
            linha["setor_responsavel"][:28],
            linha["sla_estado"],
            linha["tempo_resposta_horas"] if linha["tempo_resposta_horas"] is not None else "-",
            linha["avaliacao_nota"] if linha["avaliacao_nota"] is not None else "-",
        ])
    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabela)
    doc.build(elementos)
    arquivo.seek(0)
    return StreamingResponse(
        arquivo,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="relatorio-chamados.pdf"'},
    )

@router.get("/teste-auth")
def teste_auth(authorization: str = Header(None)):
    return {"authorization": authorization}

@router.post("/setores")
def cadastrar_setor(
    setor: SetorSchema,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
):
    _exigir_admin_ou_bootstrap(db, authorization)
    novo_setor = Setor(nome=setor.nome, sigla=setor.sigla.upper())
    db.add(novo_setor)
    db.commit()
    db.refresh(novo_setor)
    return {"mensagem": "Setor cadastrado com sucesso", "id": novo_setor.id}

@router.get("/setores")
def listar_setores(
    db: Session = Depends(get_db), 
    usuario: Usuario = Depends(get_current_user) # Qualquer usuário autenticado pode listar
):
    return db.query(Setor).all()

@router.post("/login")
def login(login: LoginSchema, db: Session = Depends(get_db)):
    from app.controllers.auth import autenticar
    from app.jwt_config import criar_access_token

    usuario = autenticar(db, login)

    if usuario is None:
        raise HTTPException(status_code=401, detail="Credenciais inválidas")

    # Cria o token com usuario_id e perfil
    access_token = criar_access_token(
        data={"sub": usuario.id, "perfil": usuario.perfil}
    )
    
    return {
        "mensagem": "Login realizado com sucesso",
        "access_token": access_token,
        "token_type": "bearer",
        "usuario_id": usuario.id,
        "perfil": usuario.perfil
    }

@router.post("/cadastrarUsuarios")
def cadastrar_usuario(
    usuarios: UsuarioCadastroSchema,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
):
    from app.models.models import Usuario
    from bcrypt import hashpw, gensalt
    _exigir_admin_ou_bootstrap(db, authorization)
    setor = db.query(Setor).filter(Setor.sigla == usuarios.setor_sigla.upper()).first()
    if setor is None:
        raise HTTPException(status_code=404, detail="Setor não encontrado")
    senha_hashed = hashpw(usuarios.senha.encode('utf-8'), gensalt()).decode('utf-8')
    usuario = Usuario(
        nome=usuarios.nome,
        email=usuarios.email,
        senha_hash=senha_hashed,
        setor_id=setor.id,
        perfil=usuarios.perfil,
    )
    db.add(usuario)
    db.commit()
    db.refresh(usuario)
    return {"mensagem": "Usuário cadastrado com sucesso", "id": usuario.id}

@router.post("/realizarChamado")
def criar_chamado(
    pedido: PedidoSchema, 
    usuario: Usuario = Depends(get_current_user), 
    db: Session = Depends(get_db)
):
    return abrirChamado(usuario, pedido, db)

@router.get("/setores/{setor_id}/usuarios")
def listar_usuarios_por_setor(
    setor_id: int, 
    db: Session = Depends(get_db), 
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    from app.models.models import Usuario
    
    # Busca apenas os usuários ativos que pertencem ao setor informado
    usuarios = db.query(Usuario).filter(
        Usuario.setor_id == setor_id, 
        Usuario.ativo == True
    ).all()
    
    # Retorna uma lista contendo apenas id e nome para o select do frontend
    return [{"id": u.id, "nome": u.nome} for u in usuarios]



@router.get("/meus-chamados")
def listar_meus_chamados(
    status: str = None,
    page: Optional[int] = None,
    per_page: int = 10,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(get_current_user)
):
    query = db.query(Chamado).filter(Chamado.usuario_solicitante_id == usuario.id)

    if status:
        query = query.filter(Chamado.status.in_(_status_equivalentes(status)))

    total = query.count()
    query = query.order_by(Chamado.data_criacao.desc())
    if page is not None:
        page = max(1, page)
        per_page = min(max(1, per_page), 100)
        query = query.offset((page - 1) * per_page).limit(per_page)

    resultado = [_serializar_chamado_lista(c, db) for c in query.all()]
    return _resposta_paginada(resultado, total, page, per_page)

@router.get("/usuarios")
def listar_usuarios(
    db: Session = Depends(get_db), 
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    # Bloqueia se for solicitante comum
    if usuario_autenticado.perfil not in ["Gestor", "Administrador"]:
        raise HTTPException(status_code=403, detail="Acesso não autorizado")

    query = db.query(Usuario)

    # REGRA DE ESCOPO: Se for Gestor, filtra apenas usuários do mesmo setor
    if usuario_autenticado.perfil == "Gestor":
        query = query.filter(Usuario.setor_id == usuario_autenticado.setor_id)
    
    usuarios = query.all()

    # Formata a resposta trazendo a sigla do setor para exibição amigável no front
    resultado = []
    for u in usuarios:
        setor = db.query(Setor).filter(Setor.id == u.setor_id).first()
        resultado.append({
            "id": u.id,
            "nome": u.nome,
            "email": u.email,
            "perfil": u.perfil,
            "setor_id": u.setor_id,
            "setor_sigla": setor.sigla if setor else "",
            "ativo": u.ativo is not False
        })
    return resultado

@router.get("/usuarios/perfil/me")
def obter_perfil_logado(
    db: Session = Depends(get_db), 
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    from app.models.models import Setor
    setor = db.query(Setor).filter(Setor.id == usuario_autenticado.setor_id).first()
    
    return {
        "nome": usuario_autenticado.nome,
        "email": usuario_autenticado.email,
        "matricula": f"#{usuario_autenticado.id}",
        "setor": setor.nome if setor else "Não informado",
        "setor_id": usuario_autenticado.setor_id,  # Adicionado para identificação no frontend
        "perfil": usuario_autenticado.perfil
    }

@router.post("/usuarios")
def cadastrar_usuario_escopo(
    payload: UsuarioCadastroSchema, 
    db: Session = Depends(get_db), 
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    if usuario_autenticado.perfil not in ["Gestor", "Administrador"]:
        raise HTTPException(status_code=403, detail="Permissão insuficiente")

    # Busca o ID do setor passado por sigla
    setor_alvo = db.query(Setor).filter(Setor.sigla == payload.setor_sigla.upper()).first()
    if not setor_alvo:
        raise HTTPException(status_code=404, detail="Setor não encontrado")

    # REGRA DE ESCOPO: Gestor não pode cadastrar usuários fora de seu setor
    if usuario_autenticado.perfil == "Gestor" and setor_alvo.id != usuario_autenticado.setor_id:
        raise HTTPException(status_code=403, detail="Gestores só podem cadastrar usuários no seu próprio setor")

    # REGRA DE ESCOPO: Gestor não pode criar um Administrador
    if usuario_autenticado.perfil == "Gestor" and payload.perfil == "Administrador":
        raise HTTPException(status_code=403, detail="Um gestor não pode criar perfis de Administrador")

    # Lógica de hash de senha (exemplo usando bcrypt compatível com seu auth.py)
    import bcrypt
    senha_hashed = bcrypt.hashpw(payload.senha.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    novo_usuario = Usuario(
        nome=payload.nome,
        email=payload.email,
        senha_hash=senha_hashed,
        perfil=payload.perfil,
        setor_id=setor_alvo.id
    )
    
    db.add(novo_usuario)
    db.commit()
    return {"mensagem": "Usuário criado com sucesso"}


@router.put("/usuarios/{id_usuario}")
def editar_usuario_escopo(
    id_usuario: int, 
    payload: dict, # Pode mapear para um Schema específico de atualização se preferir
    db: Session = Depends(get_db), 
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    if usuario_autenticado.perfil not in ["Gestor", "Administrador"]:
        raise HTTPException(status_code=403, detail="Permissão insuficiente")

    usuario_alvo = db.query(Usuario).filter(Usuario.id == id_usuario).first()
    if not usuario_alvo:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")

    # REGRA DE ESCOPO: Gestor não pode alterar usuários de outro setor
    if usuario_autenticado.perfil == "Gestor" and usuario_alvo.setor_id != usuario_autenticado.setor_id:
        raise HTTPException(status_code=403, detail="Acesso negado às informações deste setor")

    setor_alvo = db.query(Setor).filter(Setor.sigla == payload.get("setor_sigla").upper()).first()
    if not setor_alvo:
        raise HTTPException(status_code=404, detail="Setor informado inválido")

    # REGRA DE ESCOPO: Gestor não pode mover o usuário para outro setor
    if usuario_autenticado.perfil == "Gestor" and setor_alvo.id != usuario_autenticado.setor_id:
        raise HTTPException(status_code=403, detail="Você não pode mover usuários para fora do seu setor")

    # Atualização dos campos permitidos
    usuario_alvo.nome = payload.get("nome", usuario_alvo.nome)
    usuario_alvo.email = payload.get("email", usuario_alvo.email)
    usuario_alvo.perfil = payload.get("perfil", usuario_alvo.perfil)
    usuario_alvo.setor_id = setor_alvo.id
    if "ativo" in payload:
        novo_status = bool(payload.get("ativo"))
        if usuario_alvo.id == usuario_autenticado.id and not novo_status:
            raise HTTPException(status_code=422, detail="Você não pode inativar o próprio usuário")
        usuario_alvo.ativo = novo_status

    db.commit()
    return {"mensagem": "Usuário atualizado com sucesso"}


@router.get("/setores-dashboard")
def listar_setores_dashboard(
    db: Session = Depends(get_db), 
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    if usuario_autenticado.perfil not in ["Gestor", "Administrador"]:
        raise HTTPException(status_code=403, detail="Não autorizado")

    query = db.query(Setor)
    
    # REGRA DE ESCOPO: Gestor só visualiza o próprio setor
    if usuario_autenticado.perfil == "Gestor":
        query = query.filter(Setor.id == usuario_autenticado.setor_id)
        
    setores = query.order_by(Setor.nome.asc()).all()
    
    resultado = []
    for s in setores:
        # Contagem de funcionários ativos no setor
        total_funcionarios = db.query(Usuario).filter(Usuario.setor_id == s.id, Usuario.ativo == True).count()
        
        # Agregação de chamados por status vinculados ao setor_responsavel_id
        chamados_abertos = db.query(Chamado).filter(Chamado.setor_responsavel_id == s.id, Chamado.status.in_(_status_equivalentes("Aberto"))).count()
        chamados_andamento = db.query(Chamado).filter(Chamado.setor_responsavel_id == s.id, Chamado.status.in_(_status_equivalentes("Em Atendimento"))).count()
        chamados_pausados = db.query(Chamado).filter(Chamado.setor_responsavel_id == s.id, Chamado.status.in_(_status_equivalentes("Pausado"))).count()
        chamados_fechados = db.query(Chamado).filter(Chamado.setor_responsavel_id == s.id, Chamado.status.in_(_status_equivalentes("Concluído"))).count()
        
        resultado.append({
            "id": s.id,
            "nome": s.nome,
            "sigla": s.sigla,
            "total_funcionarios": total_funcionarios,
            "chamados_abertos": chamados_abertos,
            "chamados_andamento": chamados_andamento,
            "chamados_pausados": chamados_pausados,
            "chamados_fechados": chamados_fechados
        })
        
    return resultado

@router.get("/setores/{setor_id}/chamados")
def listar_chamados_do_setor(
    setor_id: int,
    page: Optional[int] = None,
    per_page: int = 10,
    db: Session = Depends(get_db),
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    if usuario_autenticado.perfil not in ["Gestor", "Administrador"]:
        raise HTTPException(status_code=403, detail="Não autorizado")
        
    # REGRA DE ESCOPO: Gestor não pode ver chamados de outros setores
    if usuario_autenticado.perfil == "Gestor" and usuario_autenticado.setor_id != setor_id:
        raise HTTPException(status_code=403, detail="Permissão negada ao escopo do setor")

    # Retorna os chamados ordenados de forma ascendente pela data_criacao (tempo de abertura)
    query = db.query(Chamado).filter(Chamado.setor_responsavel_id == setor_id).order_by(Chamado.data_criacao.asc())
    total = query.count()
    if page is not None:
        page = max(1, page)
        per_page = min(max(1, per_page), 100)
        query = query.offset((page - 1) * per_page).limit(per_page)

    chamados = [_serializar_chamado_lista(c, db) for c in query.all()]
    return _resposta_paginada(chamados, total, page, per_page)

@router.get("/setores/{setor_id}/chamados-dashboard")
def listar_chamados_dashboard(
    setor_id: int,
    db: Session = Depends(get_db),
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    from sqlalchemy import or_

    # Garante isolamento para Gestores não visualizarem dados de outros setores
    if usuario_autenticado.perfil == "Gestor" and usuario_autenticado.setor_id != setor_id:
        raise HTTPException(status_code=403, detail="Permissão negada ao escopo do setor")

    # Retorna unicamente chamados sem dono OU atribuídos diretamente ao usuário logado
    chamados = db.query(Chamado).filter(
        Chamado.setor_responsavel_id == setor_id,
        or_(
            Chamado.usuario_responsavel_id == usuario_autenticado.id,
            Chamado.usuario_responsavel_id == None
        )
    ).order_by(Chamado.data_criacao.asc()).all()
    return [serializar_chamado(c, db, True) for c in chamados]


@router.get("/relatorios/gerencial")
def relatorio_gerencial(
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    setor_id: Optional[int] = None,
    formato: str = "json",
    db: Session = Depends(get_db),
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    if usuario_autenticado.perfil not in ["Gestor", "Administrador"]:
        raise HTTPException(status_code=403, detail="Não autorizado")

    inicio = _parse_data(data_inicio)
    fim = _parse_data(data_fim, fim_do_dia=True)
    query = db.query(Chamado)

    if inicio:
        query = query.filter(Chamado.data_criacao >= inicio)
    if fim:
        query = query.filter(Chamado.data_criacao <= fim)

    if usuario_autenticado.perfil == "Gestor":
        query = query.filter(Chamado.setor_responsavel_id == usuario_autenticado.setor_id)
    elif setor_id:
        query = query.filter(Chamado.setor_responsavel_id == setor_id)

    chamados = query.order_by(Chamado.data_criacao.desc()).all()
    linhas, resumo = _linhas_relatorio(chamados, db)
    formato_normalizado = formato.lower()

    if formato_normalizado == "xlsx":
        return _resposta_xlsx(linhas, resumo)
    if formato_normalizado == "pdf":
        return _resposta_pdf(linhas, resumo)
    if formato_normalizado != "json":
        raise HTTPException(status_code=422, detail="Formato inválido. Use json, xlsx ou pdf")

    return {"resumo": resumo, "items": linhas}

@router.put("/setores/{setor_id}")
def atualizar_setor(
    setor_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    # REGRA DE SEGURANÇA: Apenas administradores alteram setores
    if usuario_autenticado.perfil != "Administrador":
        raise HTTPException(status_code=403, detail="Apenas administradores podem gerenciar setores")
        
    setor = db.query(Setor).filter(Setor.id == setor_id).first()
    if not setor:
        raise HTTPException(status_code=404, detail="Setor não encontrado")
        
    setor.nome = payload.get("nome", setor.nome)
    setor.sigla = payload.get("sigla", setor.sigla).upper()
    db.commit()
    return {"mensagem": "Setor atualizado com sucesso"}




@router.put("/usuarios/me/perfil")
def atualizar_nome_perfil(
    payload: AtualizarPerfilSchema,
    db: Session = Depends(get_db),
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    usuario_autenticado.nome = payload.nome
    db.commit()
    return {"mensagem": "Informações pessoais atualizadas com sucesso!"}

@router.put("/usuarios/me/senha")
def alterar_senha_perfil(
    payload: AlterarSenhaSchema,
    db: Session = Depends(get_db),
    usuario_autenticado: Usuario = Depends(get_current_user)
):
    from app.controllers.auth import verificarSenha
    from bcrypt import hashpw, gensalt

    if not verificarSenha(payload.senha_atual, usuario_autenticado.senha_hash):
        raise HTTPException(status_code=400, detail="A senha atual está incorreta.")

    if payload.senha_atual == payload.nova_senha:
        raise HTTPException(status_code=400, detail="A nova senha não pode ser igual à senha atual.")

    usuario_autenticado.senha_hash = hashpw(payload.nova_senha.encode('utf-8'), gensalt()).decode('utf-8')
    db.commit()
    return {"mensagem": "Senha atualizada com sucesso!"}
